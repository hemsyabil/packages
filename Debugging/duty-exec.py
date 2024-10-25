from collections import OrderedDict, defaultdict
from openpyxl import load_workbook
from datetime import datetime, timedelta
from itertools import islice
import csv
import json
from datetime import datetime
import requests


"""Module documentation here."""
__application__ = "CBSA Duty Executive Lookahead"
__author__ = "CBSA - Change Management"
__copyright__ = "Canada Border Service Agency (CBSA) - CM @ 2024"
__version__ = "1.0.0"
__maintainer__ = "Hem Raj Regmi"
__email__ = "hemraj.regmi@cbsa-asfc.gc.ca"
__status__ = "Development"

# Printing metadata
print("----------------------------------------------------------------")
print(f"Application: {__application__}")
print(f"Author: {__author__}")
print(f"Copyright: {__copyright__}")
print(f"Version: {__version__}")
print(f"Maintainer: {__maintainer__}")
print(f"Email: {__email__}")
print(f"Dev. Status: {__status__}")
print("----------------------------------------------------------------")


def translate_to_french(text):
    url = "https://libretranslate.com/translate"
    params = {
        'q': text,
        'source': 'auto',  # Auto-detect source language
        'target': 'fr',
        'format': 'text'
    }

    response = requests.post(url, data=params)

    if response.status_code == 200:
        translation = response.json()
        return translation['translatedText']

def get_day_of_week(date_string):
    date_obj = format_date_time_object(date_string)
    day_of_week = date_obj.weekday()
    days = ["Monday", "Tuesday", "Wednesday",
            "Thursday", "Friday", "Saturday", "Sunday"]
    return days[day_of_week]


def get_full_date_for_display(date_day, date_string):
    return f'{date_day}, {date_string} | {translate_to_french(date_day)}, {translate_to_french(date_string)}'


def get_upcoming_weekend():
    current_date = datetime.now().date()
    current_day_of_week = current_date.weekday()
    days_until_saturday = (5 - current_day_of_week + 7) % 7

    start_of_weekend = current_date + timedelta(days=days_until_saturday)
    end_of_weekend = start_of_weekend + timedelta(days=1)
    next_week_friday = end_of_weekend + timedelta(days=5)

    return start_of_weekend, end_of_weekend, next_week_friday


def format_date_time_object(date_time_string):
    try:
        date_time_object = datetime.strptime(
            date_time_string, "%Y-%m-%d %H:%M")
    except ValueError:
        date_time_object = datetime.strptime(date_time_string, "%Y-%m-%d")

    if isinstance(date_time_object, datetime):
        date_time_object = date_time_object.date()

    return date_time_object


def format_time_range(start_time, end_time):
    formatted_start_time = start_time.strftime('%A, %d %B %H:%M')
    formatted_end_time = end_time.strftime('%H:%M')

    duration = end_time - start_time
    hours, remainder = divmod(duration.seconds, 3600)
    minutes = remainder // 60

    return f'{formatted_start_time} to {formatted_end_time} ({hours} hours {minutes} minutes)'


def read_source_data(osr_file_path, osr_sheet_name, release_file_path):
    workbook = load_workbook(osr_file_path)
    worksheet = workbook[osr_sheet_name]
    start_of_weekend, _, next_week_friday = get_upcoming_weekend()

    maintenance_data, release_data = [], []

    for row in islice(worksheet.iter_rows(values_only=True), 1, None):
        try:
            if start_of_weekend <= row[1].date() <= next_week_friday:
                maintenance_data.append(row)
        except:
            pass

    with open(release_file_path, 'r', newline='') as csvfile:
        csv_reader = csv.reader(csvfile)

        rows_without_header = islice(csv_reader, 1, None)

        for row in rows_without_header:
            date_time_object = format_date_time_object(row[2])

            if start_of_weekend <= date_time_object <= next_week_friday:
                release_data.append(row)

    workbook.close()

    return maintenance_data, release_data


# Functions
def generate_html(maintenance_schedules, start_of_weekend, next_week_friday, total_records):

    from_date_en = start_of_weekend.strftime("%A, %B %d, %Y")
    from_date_fr = translate_to_french(from_date_en)
    to_date_en = next_week_friday.strftime("%A, %B %d, %Y")
    to_date_fr = translate_to_french(to_date_en)

    html = f"""
    <!DOCTYPE html>
    <html lang="en">

    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>CBSA Technical Change Management</title>

        <!-- styles -->
        <style>
            /* Google Font Link */
        @import url('https://fonts.googleapis.com/css2?family=Nunito+Sans:wght@200;300;400;500;600;700&display=swap');

        :root{{
            --jet-color: #333333;
            --turquoise-color: #48E5C2;
            --seasalt-color: #FCFAF9;
            --desert-sand-color: #F3D3BD;
            --davys-gray-color: #5E5E5E;
            --highlight-color: #F66247;
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
            font-family: "Nunito Sans", sans-serif;
            font-size: 12px;
            padding: 0px;
            margin: 0px;
        }}

        .banner-large {{
            padding: 20px 30px;
        }}

        .banner {{
            display: flex;
            width: 100%;
            flex-direction: column;
            align-self: flex-start;
            box-shadow: rgba(0, 0, 0, 0.15) 0px 15px 25px, rgba(0, 0, 0, 0.05) 0px 5px 10px;
        }}
        .banner .heading {{
            flex-grow: 1;
            padding-bottom: 10px;
            font-weight: 600;
            font-size: 24px;
        }}
        .banner .event-date {{
            font-weight: 600;
            color: var(--jet-color);
        }}
        .banner .event-date-mobile {{
            display: none;
        }}
        .banner .event-date .event-date-bg {{
            background-color: var(--turquoise-color);
            padding: 5px 10px;
            border-radius: 5px;
            font-weight: 500;
            line-height: 24px;
        }}

        .key-index {{
            display: flex;
            flex-direction: row;
            padding: 20px 30px;
            justify-content: space-between;
        }}
        .key-index-mobile {{
            display: none;
        }}
        .key-index .total-record {{
            font-size: 14px;
            align-self: center;
        }}
        .key-index .total-record .total-records-count {{
            background-color: var(--highlight-color);
            padding: 5px 10px;
            border-radius: 5px;
            font-weight: 600;
        }}
        .key-index .color-legend {{
            font-size: 14px;
            line-height: 30px;
            align-self: center;
        }}

        .key-index .color-legend .weekend-color, .key-index .color-legend .weekday-color {{
            display: inline-block;
            height: 30px;
            width: 40px;
            border-radius: 5px;
            margin-left: 20px;
            margin-bottom: -5px;
            box-shadow: rgba(0, 0, 0, 0.15) 0px 15px 25px, rgba(0, 0, 0, 0.05) 0px 5px 10px;
        }}
        .key-index .color-legend .weekend-color {{
            background-color: none;
        }}
        .key-index .color-legend .weekday-color {{
            background-color: var(--desert-sand-color);
        }}

        .notice {{
            display: flex;
            padding: 0px 30px;
            text-decoration: underline;
            font-weight: 500;
        }}

        .blocks-wrapped-date {{
            padding: 20px 30px;
        }}
        .blocks-wrapped-date .blocks-date {{
            background-color: var(--turquoise-color);
            padding: 5px;
            border-radius: 5px;
            font-weight: 500;
            line-height: 20px;
        }}
        .blocks-wrapped-date .blocks-all {{
            margin-top: 5px;
            display: flex;
            flex-wrap: wrap;
        }}
        .blocks-wrapped-date .blocks-all .weekday {{
            background-color: var(--desert-sand-color);
        }}
        .blocks-wrapped-date .blocks-all .block-item {{
            width: 100%;
            max-width: 280px;
            height: auto;
            box-shadow: rgba(0, 0, 0, 0.1) 0px 0px 5px 0px, rgba(0, 0, 0, 0.1) 0px 0px 1px 0px;
            border-radius: 5px;
            padding: 10px;
            margin-bottom: 20px;
            margin-right: 20px;
            cursor: pointer;
        }}
        .blocks-wrapped-date .blocks-all .block-item:hover {{
            box-shadow: rgba(0, 0, 0, 0.07) 0px 1px 2px, rgba(0, 0, 0, 0.07) 0px 2px 4px, rgba(0, 0, 0, 0.07) 0px 4px 8px, rgba(0, 0, 0, 0.07) 0px 8px 16px, rgba(0, 0, 0, 0.07) 0px 16px 32px, rgba(0, 0, 0, 0.07) 0px 32px 64px;
            transform: scale(1.1);
        }}
        .blocks-wrapped-date .blocks-all .block-item .item-horizontal-line {{
            color: var(--highlight-color);
            margin-top: 5px;
            opacity: 20%;
        }}
        .blocks-wrapped-date .blocks-all .block-item .impact-summary-show {{
            padding-top: 5px;
            font-weight: 600;
        }}
        .blocks-wrapped-date .blocks-all .block-item .impact-summary-show > span{{
            font-weight: 400;
        }}


        .popup {{
            display: none;
            position: fixed;
            top: 0px;
            left: 0px;
            width: 100%;
            height: 100%;
            overflow: auto;
            background-color: rgba(0, 0, 0, 0.5);
            justify-content: center;
        }}

        .popup  .close-button {{
            float: right;
            margin-top: 3px;
            margin-right: 3px;
            height: 40px;
            width: 40px;
            background-color: var(--highlight-color);
            text-align: center;
            line-height: 40px;
            border-radius: 10px;
            font-size: 30px;
            cursor: pointer;
        }}

        .popup .popup-content {{
            background-color: white;
            padding: 20px;
            border-radius: 5px;
            text-align: left;
            margin: 10% auto;
            width: 80%;
            position: relative;
            box-shadow: rgba(0, 0, 0, 0.1) 0px 0px 5px 0px, rgba(0, 0, 0, 0.1) 0px 0px 1px 0px;
            padding: 20px;
        }}
        .popup .popup-content .popup-content-box ul {{
            list-style: none;
        }}
        .popup .popup-content .popup-content-box .details-header {{
            background-color: var(--turquoise-color);
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 5px;
            font-size: 20px;
            font-weight: 600;
        }}
        .popup .popup-content .popup-content-box .duty-exec-list {{
            font-weight: 600;
            color: var(--jet-color);
        }}
        .popup .popup-content .popup-content-box .duty-exec-list span {{
            font-weight: 400;
        }}




        .footer {{
            display: block;
            height: auto;
            text-align: center;
            padding: 10px;
            margin-bottom: 30px;
        }}

        /* mobile layout */
        @media (max-width: 480px) {{
            .key-index {{
                flex-direction: column;
            }}
            .blocks-wrapped-date .blocks-all .block-item {{
                max-width: none;
                margin-right: 0px;
            }}
            .key-index {{
                flex-direction: column;
                align-self: flex-start;
            }}
            .key-index .total-record {{
                align-self: flex-start;
            }}
            .key-index .total-record .total-records-count {{
                padding: 5px;
            }}
            .key-index .color-legend {{
                align-self: flex-start;
            }}
            .key-index, .event-date {{
                display: none;
            }}
            .banner .event-date-mobile {{
                display: contents;
                margin-bottom: -10px;
            }}
            .banner .event-date-mobile .event-date-bg {{
                background-color: var(--turquoise-color);
                padding: 0px 10px;
                border-radius: 5px;
                font-weight: 500;
                line-height: 24px;
            }}
            .key-index-mobile {{
                display: block;
                flex-direction: row;
                padding: 20px 30px;
                justify-content: space-between;
            }}
            .key-index-mobile .total-record {{
                font-size: 14px;
                align-self: center;
            }}
            .key-index-mobile .total-record .total-records-count {{
                background-color: var(--highlight-color);
                padding: 5px 10px;
                border-radius: 5px;
                font-weight: 600;
            }}
            .key-index-mobile .color-legend {{
                font-size: 12px;
                line-height: 30px;
                align-self: center;
                display: flex;
                flex-direction: column;
            }}
            .key-index-mobile .color-legend .weekend, .key-index-mobile .color-legend .weekday {{
                align-self: flex-start;
                align-content: center;
            }}
            .key-index-mobile .color-legend .weekend span, .key-index-mobile .color-legend .weekday span {{
                display: inline-block;
                width: 60px;
                height: 30px;
                border-radius: 5px;
                margin-left: 30px;
                margin-bottom: -5px;
                box-shadow: rgba(0, 0, 0, 0.15) 0px 15px 25px, rgba(0, 0, 0, 0.05) 0px 5px 10px;
            }}
            .key-index-mobile .color-legend .weekend span {{
                background-color: none;
            }}
            .key-index-mobile .color-legend .weekday span  {{
                background-color: var(--desert-sand-color);
                margin-left: 66px;
            }}
            .footer {{
                padding: 0px 20px;
            }}
        }}
        </style>
        <link href="https://unpkg.com/boxicons@2.0.7/css/boxicons.min.css" rel="stylesheet" />
    </head>

    <body>
        <!-- Banner section -->
        <div class="banner banner-large">
            <div class="heading">
                CBSA TCM: Duty Executive Lookahead (IT Planned Outages Report) Summary <br />
                GCT de l'ASFC: Résumé de l'analyse prospective pour le directeur de service (rapport sur les pannes informatiques planifiées)
            </div>
            <div class="event-date">From <span class="event-date-bg">{from_date_en} | {from_date_fr}</span> To
                <span class="event-date-bg">{to_date_en} | {to_date_fr}</span>
            </div>

            <!-- for mobile -->
            <div class="event-date-mobile">From <span class="event-date-bg">{from_date_en} | {from_date_fr}</span> To
                <span class="event-date-bg">{to_date_en} | {to_date_fr}</span>
            </div>
        </div>

        <!-- Record index and legends -->
        <div class="key-index">
            <div class="total-record">Total Records | Total des enregistrements: 
                <span class="total-records-count">{total_records}</span>
            </div>
            <div class="color-legend">Color Legend | Légende des couleurs: 
                <span class="weekend-color"></span> Weekend | Fin de semaine
                <span class="weekday-color"></span> Weekday | Semaine
            </div>
        </div>

        <!-- Record index and legends mobile version-->
        <div class="key-index-mobile">
            <div class="total-record">Total Records | Total d’enregistrements: 
                <span class="total-records-count">{total_records}</span>
            </div>
        </div>

        <!-- Information -->
        <div class="notice">Note: Selecting any box will display the corresponding details in the pop-up box. <br>
            Remarque : La sélection d’une boîte affichera les détails correspondants dans la fenêtre contextuelle.</div>

        <!-- Display blocks -->
        <div class="blocks-wrapped-date">
    """

    loop_index = 1
    outage_summary_str = f'Impact Summary | {translate_to_french("Impact Summary")}'

    for date, outages in maintenance_schedules.items():
        week_day = get_day_of_week(date)

        html += f"""
            <div class="blocks-date">{get_full_date_for_display(week_day, date)}</div>
                <div class="blocks-all">
        """

        for outage in outages:
            start_date = outage['start_date']
            is_weekend = start_date.weekday() >= 5

            html += f"""
                    <div class="block-item {'' if is_weekend else 'weekday'}" id="duty-exec-{loop_index}">
                        <p>{outage['name_en']}</p>
                        <p>{outage['name_fr']}</p>
                        <hr class="item-horizontal-line">
                        <p class="impact-summary-show">{outage_summary_str}: <br>
                            <span>{outage['impact_en']}</span> <br>
                            <span>{outage['impact_fr']}</span>
                        </p>
                    </div>
            """
            loop_index += 1
        
        html += """</div>"""

    html += """
            </div>
            <div class="popup" id="popup">
                <div class="popup-content">
                    <span class="close-button" id="closeButton">&times;</span>
                    <div class="popup-content-box" id="popup-content-box">
                        
                    </div>
                </div>
            </div>    
        </section>

        <div class="footer">For more details please contact Technical Change Management <br> 
            Pour plus de détails, veuillez contacter la Gestion des Changements Techniques
        </div>

        <!-- scripts -->
        <script>
            var osrData = """ + json.dumps(sorted_maintenance_schedules, default=datetime_encoder) + """;

            const closeButton = document.getElementById('closeButton');
            const popup = document.getElementById('popup');

            // Function to get values based on index
            function getValuesByIndex(data, index) {
                const keys = Object.keys(data);
                if (index < 0 || index >= keys.length) {
                    return null; // Index out of bounds
                }
                const key = keys[index];
                return data[key];
            }

            document.addEventListener('DOMContentLoaded', function () {
                console.log(osrData);

                const items = document.querySelectorAll(".block-item");
                // console.log(items);

                // Flatten osrData into a single array for easier access
                const flattenedData = [];
                Object.keys(osrData).forEach(date => {
                    osrData[date].forEach(entry => {
                        flattenedData.push(entry);
                    });
                });

                items.forEach((item, index) => {
                    item.addEventListener("click", function() {
                        console.log(`Item ${index + 1} clicked`);

                        // Get the corresponding entry from the flattened data
                        const entry = flattenedData[index];

                        if (entry) {
                            // Display the entry in the display-container
                            popup.style.display = 'block';

                            document.getElementById('popup-content-box').innerHTML = `
                                <ul>
                                    <p class="details-header">Details:</p>
                                    <li class="duty-exec-list">${entry['name_en'] || ''}</li>
                                    <li class="duty-exec-list">${entry['name_fr'] || ''}</li>
                                    <br>
                                    
                                    <li class="duty-exec-list">Schedule</li>
                                    <li class="duty-exec-list">Start Date and Time: 
                                        <span>${entry['start_date'] || ''}</span>
                                    </li>
                                    <li class="duty-exec-list">End Date and Time: 
                                        <span>${entry['end_date'] || ''}</span>
                                    </li>
                                    <br>
                                    
                                    <li class="duty-exec-list">Outage</li>
                                    <li class="duty-exec-list">
                                        <span>${entry['outage_en'] || ''}</span>
                                    </li>
                                    <li class="duty-exec-list">
                                        <span>${entry['outage_fr'] || ''}</span>
                                    </li>
                                    <br>
                                    
                                    <li class="duty-exec-list">Impact Summary: <br>
                                        <span>${entry['impact_en'] || ''}</span> <br>
                                        <span>${entry['impact_fr'] || ''}</span>
                                    </li>
                                </ul>`;
                        } else {
                            console.log('No entry found for index:', index);
                        }
                    });
                });
            });
            
            closeButton.addEventListener('click', () => {
                popup.style.display = 'none';
            });
                

        </script>
    </body>

    </html>
    """

    return html


if __name__ == "__main__":
    osr_file_path = "assets/OSR_Data.xlsm"
    release_file_path = "assets/CM Release Dates  IT Service Management.csv"
    osr_sheet_name = "OSR_Data"

    start_of_weekend, end_of_weekend, next_week_friday = get_upcoming_weekend()

    maintenance_data, release_data = read_source_data(
        osr_file_path, osr_sheet_name, release_file_path)

    maintenance_schedules = {}

    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d %H:%M")
        except ValueError:
            return datetime.strptime(date_str, "%Y-%m-%d")

    for entry in maintenance_data:
        start_date = entry[0].strftime('%Y-%m-%d')
        if start_date not in maintenance_schedules:
            maintenance_schedules[start_date] = []

        maintenance_schedules[start_date].append({
            'name_en': entry[3],
            'name_fr': entry[6],
            'start_date': entry[0],
            'end_date': entry[1],
            'outage_en': entry[4],
            'outage_fr': entry[7],
            'impact_en': entry[8],
            'impact_fr': entry[9]
        })

    for entry in release_data:
        start_date = parse_date(entry[2]).strftime('%Y-%m-%d')
        if start_date not in maintenance_schedules:
            maintenance_schedules[start_date] = []

        maintenance_schedules[start_date].append({
            'name_en': entry[0],
            'name_fr': '',
            'start_date': parse_date(entry[2]),
            'end_date': parse_date(entry[3]),
            'outage_en': entry[4],
            'outage_fr': entry[5],
            'impact_en': entry[4],
            'impact_fr': ''
        })

    total_records = len(maintenance_data) + len(release_data)

    # Sort the keys
    sorted_keys = sorted(maintenance_schedules.keys())

    # Create a new OrderedDict with sorted keys
    sorted_maintenance_schedules = OrderedDict()
    for key in sorted_keys:
        sorted_maintenance_schedules[key] = maintenance_schedules[key]

    def datetime_encoder(obj):
        if isinstance(obj, datetime):
            return obj.isoformat()

    html_content = generate_html(
        sorted_maintenance_schedules, start_of_weekend, next_week_friday, total_records)
    file_name = f'Duty Executive Lookahead: {start_of_weekend} - {next_week_friday}.html'
    with open(file_name, "w") as f:
        f.write(html_content)
